from openpyxl import load_workbook


class Recipe:
    def __init__(self, name, author, recipe_type, description, link, ingredient, kitchen_country):
        self.name = name
        self.author = author
        self.recipe_type = recipe_type
        self.description = description
        self.link = link
        self.ingredient = ingredient
        self.kitchen_country = kitchen_country

    def __str__(self):
        return ",".join([f"{key}:{str(value)}" for key, value in self.__dict__.items()])


class RecipeNotFoundError(Exception):
    pass


class Model:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname
        self.wb = load_workbook(filename)
        self.ws = self.wb[sheetname]
        ws_values = list(self.ws.values)
        self.header = ws_values[0]
        self.database = [self.__obj_from_row__(row_record) for row_record in ws_values[1:]]

    @staticmethod
    def __obj_from_row__(row_record):
        data_from_row = [elm if elm else "" for elm in row_record]
        return Recipe(*data_from_row)

    @staticmethod
    def __obj_to_row__(obj):
        return [elm for elm in obj.__dict__.values()]

    @staticmethod
    def get_one_record_by(database, word):
        data = []
        for rec in database:
            if word in "".join(str(rec.__dict__.values())):
                data.append(rec)
        return data

    def save_db_to_xlsx(self, filename, reinit=True):
        self.wb.save(filename)
        if reinit:
            self.__init__(self.filename, self.sheetname)

    def delete_rec(self, num_rec):
        self.database.pop(num_rec - 1)
        self.delete_from_xlsx(num_rec)

    def delete_from_xlsx(self, num_rec):
        self.ws.delete_rows(num_rec + 1, 1)
        self.save_db_to_xlsx(self.filename, False)

    def change_rec_obj(self, num_rec, new_row_data):
        changing_obj = self.database[num_rec - 1]
        for col, elm in enumerate(changing_obj.__dict__):
            changing_obj.__setattr__(elm, new_row_data[col])
        self.change_rec_xlsx(num_rec + 1, new_row_data)

    def change_rec_xlsx(self, num_rec, new_row_data):
        for col, elm in enumerate(self.header):
            self.ws.cell(row=num_rec, column=col + 1).value = new_row_data[col]
        self.save_db_to_xlsx(self.filename, False)

    def add_new_obj(self, num_rec, new_row):
        self.database.append(self.__obj_from_row__(new_row))
        self.change_rec_xlsx(num_rec, new_row)
